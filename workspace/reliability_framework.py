#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
10kV配电线路供电可靠性计算泛化框架
从参数文件读取：常量、输入路径/Sheet、字段映射、输出路径；
支持任意线路 Excel，按技术方案计算并输出结果。
用法: python reliability_framework.py [参数文件路径]
默认参数文件: config/reliability_params.json
"""

import json
import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def load_config(config_path):
    """从 JSON 文件加载参数。"""
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg


def _log(msg, verbose):
    if verbose:
        print(msg)


def parse_laying_weights_and_fault_rate(line_model, constants):
    """
    敷设方式解析：带JK→架空，None→忽略，不带JK→电缆。
    返回: (电缆权重, 架空权重, 加权故障率, 描述)
    """
    s = str(line_model).strip()
    segments = [x.strip() for x in s.replace("\r", "\n").split("\n") if x.strip()]
    cable_w = 0.0
    overhead_w = 0.0
    for seg in segments:
        if ":" not in seg:
            continue
        parts = seg.rsplit(":", 1)
        name_raw = parts[0].strip()
        name = name_raw.upper()
        if name == "NONE" or not name_raw:
            continue
        try:
            w = float(parts[1].strip().rstrip("%").strip()) / 100.0
        except Exception:
            continue
        if "JK" in name:
            overhead_w += w
        else:
            cable_w += w
    total = cable_w + overhead_w
    if total <= 0:
        cable_w, overhead_w = 0.0, 1.0
        total = 1.0
    rate = (cable_w * constants["Cable_Fault_Rate"] + overhead_w * constants["Overhead_Fault_Rate"]) / total
    desc = f"电缆{cable_w/total*100:.1f}%+架空{overhead_w/total*100:.1f}%"
    return cable_w / total, overhead_w / total, rate, desc


def get_isolation_time(auto_status, constants):
    if isinstance(auto_status, bool):
        return constants["Auto_Isolation_Time"] if auto_status else constants["Manual_Isolation_Time"]
    return constants["Auto_Isolation_Time"] if str(auto_status).upper() == "TRUE" else constants["Manual_Isolation_Time"]


def clean_data(df, line_type, verbose):
    original_count = len(df)
    df["长度(km)"] = pd.to_numeric(df["长度(km)"], errors="coerce")
    df["用户数(台)"] = pd.to_numeric(df["用户数(台)"], errors="coerce")
    df = df[(df["长度(km)"] >= 0) & (df["用户数(台)"] >= 0)].dropna(subset=["长度(km)", "用户数(台)"])
    _log(f"{line_type}: 原始{original_count}行 → 清洗后{len(df)}行", verbose)
    return df


def calculate_segment_indicators(df, line_total_users, line_type, constants, verbose):
    df = df.copy()
    df["有效分段"] = df["用户数(台)"] > 0
    df["故障次数(次/年)"] = np.where(df["有效分段"], df["长度(km)"] * df["故障率"], 0)
    df["故障总时间(小时/次)"] = df["隔离时间"] + constants["Cable_Repair_Time"]
    df["SAIDI-F"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["故障次数(次/年)"] * df["故障总时间(小时/次)"] * df["用户数(台)"]) / line_total_users,
        0,
    )
    df["SAIFI-F"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["故障次数(次/年)"] * df["用户数(台)"]) / line_total_users,
        0,
    )
    df["预安排次数(次/年)"] = np.where(
        df["有效分段"],
        df["长度(km)"] * constants["Scheduled_Outage_Rate"],
        0,
    )
    df["SAIDI-S"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["预安排次数(次/年)"] * constants["Scheduled_Total_Time"] * df["用户数(台)"]) / line_total_users,
        0,
    )
    df["SAIFI-S"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["预安排次数(次/年)"] * df["用户数(台)"]) / line_total_users,
        0,
    )
    df["SAIDI合计"] = df["SAIDI-F"] + df["SAIDI-S"]
    df["SAIFI合计"] = df["SAIFI-F"] + df["SAIFI-S"]

    if verbose:
        _log(f"\n--- {line_type}分段级计算（分母={line_total_users}） ---", verbose)
        for idx, row in df.iterrows():
            _log(f"  【{row['分段编号']}】 长度={row['长度(km)']}km 用户={row['用户数(台)']} 有效={row['有效分段']} 故障率={row['故障率']:.6f} SAIDI合计={row['SAIDI合计']:.6f} SAIFI合计={row['SAIFI合计']:.6f}", verbose)
    return df


def calculate_summary(df, line_total_users, line_type, constants, verbose):
    total_length = df["长度(km)"].sum()
    total_fault_count = df["故障次数(次/年)"].sum()
    total_scheduled_count = df["预安排次数(次/年)"].sum()
    saidi_f = df["SAIDI-F"].sum()
    saidi_s = df["SAIDI-S"].sum()
    saidi_total = saidi_f + saidi_s
    saifi_f = df["SAIFI-F"].sum()
    saifi_s = df["SAIFI-S"].sum()
    saifi_total = saifi_f + saifi_s
    if line_total_users > 0:
        theory_hours = line_total_users * constants["Annual_Power_Hours"]
        outage_hours = saidi_total * line_total_users
        asai = ((theory_hours - outage_hours) / theory_hours) * 100
    else:
        asai = 100.0
    if verbose:
        _log(f"\n--- {line_type}汇总 --- 总长度={total_length:.4f}km 总用户={line_total_users} SAIDI合计={saidi_total:.6f} SAIFI合计={saifi_total:.6f} ASAI={asai:.6f}%", verbose)
    return {
        "线路类型": line_type,
        "总长度(km)": round(total_length, 4),
        "总用户数(台)": line_total_users,
        "总故障次数(次/年)": round(total_fault_count, 6),
        "总预安排次数(次/年)": round(total_scheduled_count, 6),
        "SAIDI-F": round(saidi_f, 6),
        "SAIDI-S": round(saidi_s, 6),
        "SAIDI合计": round(saidi_total, 6),
        "SAIFI-F": round(saifi_f, 6),
        "SAIFI-S": round(saifi_s, 6),
        "SAIFI合计": round(saifi_total, 6),
        "ASAI(%)": round(asai, 6),
    }


def run(config_path=None):
    if config_path is None:
        base = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(base, "config", "reliability_params.json")
    config = load_config(config_path)
    constants = config["constants"]
    inp = config["input"]
    out_cfg = config.get("output", {})
    field_mappings = config["field_mappings"]
    verbose = config.get("verbose", True)

    excel_path = inp["excel_path"]
    main_sheet = inp["main_sheet"]
    branch_sheet = inp["branch_sheet"]
    output_path = out_cfg.get("excel_path") or excel_path.replace(".xlsx", "_可靠性计算结果.xlsx")

    main_map = field_mappings["main"]
    branch_map = field_mappings["branch"]

    def step(title, fn):
        if verbose:
            print("=" * 80)
            print(title)
            print("=" * 80)
        return fn()

    # 1) 常量
    step("【第一步】核心常量", lambda: [_log(f"  {k}: {v}", verbose) for k, v in constants.items()] or None)

    # 2) 读取 Excel
    def read_excel():
        xls = pd.ExcelFile(excel_path)
        _log(f"Excel: {excel_path}", verbose)
        _log(f"Sheet: {xls.sheet_names}", verbose)
        df_main = pd.read_excel(xls, main_sheet)
        df_branch = pd.read_excel(xls, branch_sheet)
        _log(f"主线行数: {len(df_main)}  分支行数: {len(df_branch)}", verbose)
        return df_main, df_branch

    df_main, df_branch = step("【第二步】读取Excel", read_excel)

    # 3) 字段映射
    def do_mapping():
        df_m = df_main.rename(columns=main_map)[list(main_map.values())]
        df_b = df_branch.rename(columns=branch_map)[list(branch_map.values())]
        _log("主线列: " + str(list(main_map.values())), verbose)
        _log("分支列: " + str(list(branch_map.values())), verbose)
        return df_m, df_b

    df_main_mapped, df_branch_mapped = step("【第三步】字段映射", do_mapping)

    # 4) 数据清洗
    df_main_clean = clean_data(df_main_mapped.copy(), "主线", verbose)
    df_branch_clean = clean_data(df_branch_mapped.copy(), "分支", verbose)

    if verbose:
        print("=" * 80)
        print("【第四步】数据清洗")
        print("=" * 80)

    # 5) 敷设方式解析 + 故障率、隔离时间
    if verbose:
        print("=" * 80)
        print("【第五步】敷设方式解析（带JK→架空，None→忽略，不带JK→电缆）")
        print("=" * 80)

    for df, name in [(df_main_clean, "主线"), (df_branch_clean, "分支")]:
        parsed = df["敷设方式_原始"].apply(lambda x: parse_laying_weights_and_fault_rate(x, constants))
        df["电缆权重"] = [x[0] for x in parsed]
        df["架空权重"] = [x[1] for x in parsed]
        df["故障率"] = [x[2] for x in parsed]
        df["敷设方式描述"] = [x[3] for x in parsed]
        df["隔离时间"] = df["自动化状态"].apply(lambda x: get_isolation_time(x, constants))
        if verbose:
            for idx, row in df.iterrows():
                _log(f"  {row['分段编号']}: {row['敷设方式描述']} 故障率={row['故障率']:.6f}", verbose)

    # 6) 线路总用户数
    if verbose:
        print("=" * 80)
        print("【第六步】线路总用户数")
        print("=" * 80)
    main_total_users = int(df_main_clean["用户数(台)"].sum())
    branch_total_users = int(df_branch_clean["用户数(台)"].sum())
    all_total_users = main_total_users + branch_total_users
    _log(f"主线={main_total_users} 分支={branch_total_users} 全线路={all_total_users}", verbose)

    # 7) 分段级指标
    if verbose:
        print("=" * 80)
        print("【第七步】分段级可靠性指标计算")
        print("=" * 80)
    df_main_result = calculate_segment_indicators(df_main_clean, main_total_users, "主线", constants, verbose)
    df_branch_result = calculate_segment_indicators(df_branch_clean, branch_total_users, "分支", constants, verbose)

    # 8) 汇总级指标
    if verbose:
        print("=" * 80)
        print("【第八步】汇总级指标")
        print("=" * 80)
    main_summary = calculate_summary(df_main_result, main_total_users, "主线", constants, verbose)
    branch_summary = calculate_summary(df_branch_result, branch_total_users, "分支", constants, verbose)

    # 9) 全线路加权汇总
    all_saidi_f = (main_summary["SAIDI-F"] * main_total_users + branch_summary["SAIDI-F"] * branch_total_users) / all_total_users
    all_saidi_s = (main_summary["SAIDI-S"] * main_total_users + branch_summary["SAIDI-S"] * branch_total_users) / all_total_users
    all_saidi_total = all_saidi_f + all_saidi_s
    all_saifi_f = (main_summary["SAIFI-F"] * main_total_users + branch_summary["SAIFI-F"] * branch_total_users) / all_total_users
    all_saifi_s = (main_summary["SAIFI-S"] * main_total_users + branch_summary["SAIFI-S"] * branch_total_users) / all_total_users
    all_saifi_total = all_saifi_f + all_saifi_s
    all_theory = all_total_users * constants["Annual_Power_Hours"]
    all_asai = ((all_theory - all_saidi_total * all_total_users) / all_theory) * 100
    all_summary = {
        "线路类型": "全线路",
        "总长度(km)": round(main_summary["总长度(km)"] + branch_summary["总长度(km)"], 4),
        "总用户数(台)": all_total_users,
        "总故障次数(次/年)": round(main_summary["总故障次数(次/年)"] + branch_summary["总故障次数(次/年)"], 6),
        "总预安排次数(次/年)": round(main_summary["总预安排次数(次/年)"] + branch_summary["总预安排次数(次/年)"], 6),
        "SAIDI-F": round(all_saidi_f, 6),
        "SAIDI-S": round(all_saidi_s, 6),
        "SAIDI合计": round(all_saidi_total, 6),
        "SAIFI-F": round(all_saifi_f, 6),
        "SAIFI-S": round(all_saifi_s, 6),
        "SAIFI合计": round(all_saifi_total, 6),
        "ASAI(%)": round(all_asai, 6),
    }

    summary_df = pd.DataFrame([main_summary, branch_summary, all_summary])
    if verbose:
        print("=" * 80)
        print("【第九步】最终汇总")
        print("=" * 80)
        print(summary_df.to_string(index=False))

    # 10) 输出 Excel
    output_cols = [
        "分段编号", "长度(km)", "用户数(台)", "电缆权重", "架空权重", "敷设方式描述", "自动化状态", "故障率", "隔离时间",
        "故障次数(次/年)", "故障总时间(小时/次)", "预安排次数(次/年)",
        "SAIDI-F", "SAIDI-S", "SAIDI合计", "SAIFI-F", "SAIFI-S", "SAIFI合计",
    ]
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet(title="主线分段明细")
    for r in dataframe_to_rows(df_main_result[output_cols], index=False, header=True):
        ws1.append(r)
    ws2 = wb.create_sheet(title="分支分段明细")
    for r in dataframe_to_rows(df_branch_result[output_cols], index=False, header=True):
        ws2.append(r)
    ws3 = wb.create_sheet(title="指标汇总")
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws3.append(r)
    wb.save(output_path)
    print(f"\n结果已保存: {output_path}")
    return summary_df, output_path


if __name__ == "__main__":
    config_path = sys.argv[1] if len(sys.argv) > 1 else None
    run(config_path)
