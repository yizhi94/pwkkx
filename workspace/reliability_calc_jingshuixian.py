#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
10kV配电线路供电可靠性计算脚本
按照《技术方案》严格实现，每步打印结果便于验证
针对: 10kV景704景水线.xlsx
"""

import pandas as pd
import numpy as np

# ==================== 第一步：定义核心常量（按技术方案） ====================
print("=" * 80)
print("【第一步】定义核心常量（按技术方案）")
print("=" * 80)

# 技术方案规定的常量（敷设方式仅保留：电缆、架空，按权重加权）
CONSTANTS = {
    "Cable_Fault_Rate": 0.09282879,      # 电缆故障率 次/(km·年)，YJV/YJLV/YJV
    "Overhead_Fault_Rate": 0.15337829,   # 架空故障率 次/(km·年)，JK 开头
    "Auto_Isolation_Time": 0.557,        # 自动化隔离时间 小时/次
    "Manual_Isolation_Time": 2.0,        # 非自动化隔离时间 小时/次
    "Cable_Repair_Time": 3.073,          # 故障修复时间 小时/次
    "Scheduled_Outage_Rate": 0.0221,     # 预安排停电率 次/(km·年)
    "Scheduled_Total_Time": 5.475,       # 预安排停电总时间 小时/次
    "Annual_Power_Hours": 8760           # 年标准供电时间 小时/年
}

for key, value in CONSTANTS.items():
    print(f"  {key}: {value}")

# ==================== 第二步：读取Excel数据 ====================
print("\n" + "=" * 80)
print("【第二步】读取Excel数据")
print("=" * 80)

excel_path = r"/mnt/d/pwkkx/document/10kV景704景水线.xlsx"
xls = pd.ExcelFile(excel_path)
print(f"Excel文件路径: {excel_path}")
print(f"Sheet列表: {xls.sheet_names}")

df_main = pd.read_excel(xls, '主线')
df_branch = pd.read_excel(xls, '分支')

print(f"\n主线数据行数: {len(df_main)}")
print(f"分支数据行数: {len(df_branch)}")

# ==================== 第三步：字段映射（按技术方案3.1.1） ====================
print("\n" + "=" * 80)
print("【第三步】字段映射（按技术方案3.1.1）")
print("=" * 80)

# 主线字段映射
main_mapping = {
    "线路分段": "分段编号",
    "起点是否自动化": "自动化状态",
    "长度(km)": "长度(km)",
    "用户数量(台)": "用户数(台)",
    "线路型号": "敷设方式_原始"
}

# 分支字段映射
branch_mapping = {
    "分支分段": "分段编号",
    "是否自动化": "自动化状态",
    "长度(km)": "长度(km)",
    "用户数量(台)": "用户数(台)",
    "线路型号": "敷设方式_原始"
}

print("主线字段映射:")
for orig, new in main_mapping.items():
    print(f"  {orig} → {new}")

print("\n分支字段映射:")
for orig, new in branch_mapping.items():
    print(f"  {orig} → {new}")

# 执行映射
df_main_mapped = df_main.rename(columns=main_mapping)[list(main_mapping.values())]
df_branch_mapped = df_branch.rename(columns=branch_mapping)[list(branch_mapping.values())]

print("\n映射后主线数据:")
print(df_main_mapped.to_string())

print("\n映射后分支数据:")
print(df_branch_mapped.to_string())

# ==================== 第四步：数据清洗（按技术方案3.1.2） ====================
print("\n" + "=" * 80)
print("【第四步】数据清洗（按技术方案3.1.2）")
print("=" * 80)

def clean_data(df, line_type):
    """数据清洗：类型转换、无效数据过滤"""
    original_count = len(df)
    
    # 转换数值类型
    df["长度(km)"] = pd.to_numeric(df["长度(km)"], errors="coerce")
    df["用户数(台)"] = pd.to_numeric(df["用户数(台)"], errors="coerce")
    
    # 过滤无效数据
    df = df[(df["长度(km)"] >= 0) & (df["用户数(台)"] >= 0)].dropna(subset=["长度(km)", "用户数(台)"])
    
    cleaned_count = len(df)
    print(f"{line_type}: 原始{original_count}行 → 清洗后{cleaned_count}行")
    
    return df

df_main_clean = clean_data(df_main_mapped.copy(), "主线")
df_branch_clean = clean_data(df_branch_mapped.copy(), "分支")

# ==================== 第五步：敷设方式解析与加权故障率（电缆 / 架空 + 权重） ====================
print("\n" + "=" * 80)
print("【第五步】敷设方式解析：带JK→架空，None→忽略，不带JK→电缆，按各自权重加权故障率")
print("=" * 80)

def parse_laying_weights_and_fault_rate(line_model):
    """
    解析线路型号，得到电缆权重、架空权重、加权故障率。
    格式示例: "PD_JKLYJ-300: 50.00%\\nPD_VLY-8.7/10-3×300: 50.00%" 或 "PD_JKLYJ: 71.43%\\nNone: 14.29%\\nPD_VLY: 14.29%"
    规则: 带 JK → 架空；None 直接忽略不做计算；不带 JK → 普通电缆。
    仅对非 None 部分按权重归一后加权：故障率 = 电缆权重×电缆故障率 + 架空权重×架空故障率。
    返回: (电缆权重, 架空权重, 加权故障率, 描述字符串)，权重为归一后的有效权重（和为1）
    """
    s = str(line_model).strip()
    segments = [x.strip() for x in s.replace("\r", "\n").split("\n") if x.strip()]
    cable_w = 0.0
    overhead_w = 0.0
    for seg in segments:
        if ":" not in seg:
            continue
        parts = seg.rsplit(":", 1)
        name_raw = parts[0].strip()
        name = name_raw.upper()
        if name == "NONE" or not name_raw:
            continue
        try:
            w = float(parts[1].strip().rstrip("%").strip()) / 100.0
        except Exception:
            continue
        if "JK" in name:
            overhead_w += w
        else:
            cable_w += w
    total = cable_w + overhead_w
    if total <= 0:
        cable_w, overhead_w = 0.0, 1.0
        total = 1.0
    rate = (cable_w * CONSTANTS["Cable_Fault_Rate"] + overhead_w * CONSTANTS["Overhead_Fault_Rate"]) / total
    desc = f"电缆{cable_w/total*100:.1f}%+架空{overhead_w/total*100:.1f}%"
    return cable_w / total, overhead_w / total, rate, desc

def apply_laying_parse(series):
    """对整列应用解析，返回 (电缆权重, 架空权重, 加权故障率, 描述) 的列表，便于赋给多列"""
    out = series.apply(parse_laying_weights_and_fault_rate)
    return out

# 主线
parsed_main = df_main_clean["敷设方式_原始"].apply(parse_laying_weights_and_fault_rate)
df_main_clean["电缆权重"] = [x[0] for x in parsed_main]
df_main_clean["架空权重"] = [x[1] for x in parsed_main]
df_main_clean["故障率"] = [x[2] for x in parsed_main]
df_main_clean["敷设方式描述"] = [x[3] for x in parsed_main]

# 分支
parsed_branch = df_branch_clean["敷设方式_原始"].apply(parse_laying_weights_and_fault_rate)
df_branch_clean["电缆权重"] = [x[0] for x in parsed_branch]
df_branch_clean["架空权重"] = [x[1] for x in parsed_branch]
df_branch_clean["故障率"] = [x[2] for x in parsed_branch]
df_branch_clean["敷设方式描述"] = [x[3] for x in parsed_branch]

print("主线敷设方式与加权故障率:")
for idx, row in df_main_clean.iterrows():
    raw = str(row["敷设方式_原始"])[:50]
    print(f"  {row['分段编号']}: {raw}...")
    print(f"      → {row['敷设方式描述']}，故障率={row['故障率']:.6f} (电缆率={CONSTANTS['Cable_Fault_Rate']}，架空率={CONSTANTS['Overhead_Fault_Rate']})")

print("\n分支敷设方式与加权故障率:")
for idx, row in df_branch_clean.iterrows():
    raw = str(row["敷设方式_原始"])[:50]
    print(f"  {row['分段编号']}: {raw}...")
    print(f"      → {row['敷设方式描述']}，故障率={row['故障率']:.6f}")

# ==================== 第六步：计算线路总用户数（按技术方案3.1.4） ====================
print("\n" + "=" * 80)
print("【第六步】计算线路总用户数（按技术方案3.1.4）")
print("=" * 80)

main_total_users = df_main_clean["用户数(台)"].sum()
branch_total_users = df_branch_clean["用户数(台)"].sum()
all_total_users = main_total_users + branch_total_users

print(f"主线总用户数 = {main_total_users} 台")
print(f"分支总用户数 = {branch_total_users} 台")
print(f"全线路总用户数 = {main_total_users} + {branch_total_users} = {all_total_users} 台")

# ==================== 第七步：匹配隔离时间（故障率已在第五步由电缆/架空权重得到） ====================
print("\n" + "=" * 80)
print("【第七步】匹配隔离时间")
print("=" * 80)

def get_isolation_time(auto_status):
    """根据自动化状态获取隔离时间"""
    if isinstance(auto_status, bool):
        return CONSTANTS["Auto_Isolation_Time"] if auto_status else CONSTANTS["Manual_Isolation_Time"]
    else:
        return CONSTANTS["Auto_Isolation_Time"] if str(auto_status).upper() == "TRUE" else CONSTANTS["Manual_Isolation_Time"]

df_main_clean["隔离时间"] = df_main_clean["自动化状态"].apply(get_isolation_time)
df_branch_clean["隔离时间"] = df_branch_clean["自动化状态"].apply(get_isolation_time)

print("主线参数（故障率已按电缆/架空权重加权）:")
print(df_main_clean[["分段编号", "电缆权重", "架空权重", "敷设方式描述", "故障率", "自动化状态", "隔离时间"]].to_string())

print("\n分支参数:")
print(df_branch_clean[["分段编号", "电缆权重", "架空权重", "敷设方式描述", "故障率", "自动化状态", "隔离时间"]].to_string())

# ==================== 第八步：分段级可靠性指标计算（按技术方案第四章） ====================
print("\n" + "=" * 80)
print("【第八步】分段级可靠性指标计算（按技术方案第四章）")
print("=" * 80)

def calculate_segment_indicators(df, line_total_users, line_type):
    """
    计算分段级指标
    按技术方案：分母使用"线路总用户数"（主线用主线总用户数，支线用支线总用户数）
    """
    print(f"\n--- {line_type}分段级计算（分母={line_total_users}） ---")
    
    # 标记有效分段（用户数 > 0）
    df["有效分段"] = df["用户数(台)"] > 0
    
    # 4.1 故障类指标
    # 故障次数 = 长度(km) × 故障率
    df["故障次数(次/年)"] = np.where(
        df["有效分段"],
        df["长度(km)"] * df["故障率"],
        0
    )
    
    # 故障总时间 = 隔离时间 + 修复时间
    df["故障总时间(小时/次)"] = df["隔离时间"] + CONSTANTS["Cable_Repair_Time"]
    
    # SAIDI-F = (故障次数 × 故障总时间 × 分段用户数) ÷ 线路总用户数
    df["SAIDI-F"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["故障次数(次/年)"] * df["故障总时间(小时/次)"] * df["用户数(台)"]) / line_total_users,
        0
    )
    
    # SAIFI-F = (故障次数 × 分段用户数) ÷ 线路总用户数
    df["SAIFI-F"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["故障次数(次/年)"] * df["用户数(台)"]) / line_total_users,
        0
    )
    
    # 4.2 预安排类指标
    # 预安排次数 = 长度(km) × 预安排停电率
    df["预安排次数(次/年)"] = np.where(
        df["有效分段"],
        df["长度(km)"] * CONSTANTS["Scheduled_Outage_Rate"],
        0
    )
    
    # SAIDI-S = (预安排次数 × 预安排总时间 × 分段用户数) ÷ 线路总用户数
    df["SAIDI-S"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["预安排次数(次/年)"] * CONSTANTS["Scheduled_Total_Time"] * df["用户数(台)"]) / line_total_users,
        0
    )
    
    # SAIFI-S = (预安排次数 × 分段用户数) ÷ 线路总用户数
    df["SAIFI-S"] = np.where(
        df["有效分段"] & (line_total_users > 0),
        (df["预安排次数(次/年)"] * df["用户数(台)"]) / line_total_users,
        0
    )
    
    # 4.3 合计类指标
    df["SAIDI合计"] = df["SAIDI-F"] + df["SAIDI-S"]
    df["SAIFI合计"] = df["SAIFI-F"] + df["SAIFI-S"]
    
    # 打印每个分段的计算过程
    for idx, row in df.iterrows():
        print(f"\n  【{row['分段编号']}】")
        print(f"    长度={row['长度(km)']}km, 用户数={row['用户数(台)']}台, 有效={row['有效分段']}")
        print(f"    敷设方式={row['敷设方式描述']}, 故障率={row['故障率']:.6f}")
        print(f"    自动化={row['自动化状态']}, 隔离时间={row['隔离时间']}h")
        print(f"    故障次数 = {row['长度(km)']} × {row['故障率']} = {row['故障次数(次/年)']:.6f} 次/年")
        print(f"    故障总时间 = {row['隔离时间']} + {CONSTANTS['Cable_Repair_Time']} = {row['故障总时间(小时/次)']:.3f} h/次")
        if row['有效分段'] and line_total_users > 0:
            print(f"    SAIDI-F = ({row['故障次数(次/年)']:.6f} × {row['故障总时间(小时/次)']:.3f} × {row['用户数(台)']}) / {line_total_users} = {row['SAIDI-F']:.6f}")
            print(f"    SAIFI-F = ({row['故障次数(次/年)']:.6f} × {row['用户数(台)']}) / {line_total_users} = {row['SAIFI-F']:.6f}")
            print(f"    预安排次数 = {row['长度(km)']} × {CONSTANTS['Scheduled_Outage_Rate']} = {row['预安排次数(次/年)']:.6f} 次/年")
            print(f"    SAIDI-S = ({row['预安排次数(次/年)']:.6f} × {CONSTANTS['Scheduled_Total_Time']} × {row['用户数(台)']}) / {line_total_users} = {row['SAIDI-S']:.6f}")
            print(f"    SAIFI-S = ({row['预安排次数(次/年)']:.6f} × {row['用户数(台)']}) / {line_total_users} = {row['SAIFI-S']:.6f}")
        print(f"    SAIDI合计 = {row['SAIDI-F']:.6f} + {row['SAIDI-S']:.6f} = {row['SAIDI合计']:.6f}")
        print(f"    SAIFI合计 = {row['SAIFI-F']:.6f} + {row['SAIFI-S']:.6f} = {row['SAIFI合计']:.6f}")
    
    return df

# 计算主线和分支的分段级指标
df_main_result = calculate_segment_indicators(df_main_clean.copy(), main_total_users, "主线")
df_branch_result = calculate_segment_indicators(df_branch_clean.copy(), branch_total_users, "分支")

# ==================== 第九步：汇总级指标计算（按技术方案第五章） ====================
print("\n" + "=" * 80)
print("【第九步】汇总级指标计算（按技术方案第五章）")
print("=" * 80)

def calculate_summary(df, line_total_users, line_type):
    """计算汇总级指标（按技术方案5.1）"""
    total_length = df["长度(km)"].sum()
    total_fault_count = df["故障次数(次/年)"].sum()
    total_scheduled_count = df["预安排次数(次/年)"].sum()
    
    saidi_f = df["SAIDI-F"].sum()
    saidi_s = df["SAIDI-S"].sum()
    saidi_total = saidi_f + saidi_s
    
    saifi_f = df["SAIFI-F"].sum()
    saifi_s = df["SAIFI-S"].sum()
    saifi_total = saifi_f + saifi_s
    
    # ASAI计算（按技术方案5.1）
    if line_total_users > 0:
        theory_hours = line_total_users * CONSTANTS["Annual_Power_Hours"]
        outage_hours = saidi_total * line_total_users
        asai = ((theory_hours - outage_hours) / theory_hours) * 100
    else:
        asai = 100.0
    
    print(f"\n--- {line_type}汇总 ---")
    print(f"  总长度 = {total_length:.4f} km")
    print(f"  总用户数 = {line_total_users} 台")
    print(f"  总故障次数 = {total_fault_count:.6f} 次/年")
    print(f"  总预安排次数 = {total_scheduled_count:.6f} 次/年")
    print(f"  SAIDI-F = {saidi_f:.6f} 小时/(户·年)")
    print(f"  SAIDI-S = {saidi_s:.6f} 小时/(户·年)")
    print(f"  SAIDI合计 = {saidi_total:.6f} 小时/(户·年)")
    print(f"  SAIFI-F = {saifi_f:.6f} 次/(户·年)")
    print(f"  SAIFI-S = {saifi_s:.6f} 次/(户·年)")
    print(f"  SAIFI合计 = {saifi_total:.6f} 次/(户·年)")
    print(f"  ASAI计算: ({line_total_users} × {CONSTANTS['Annual_Power_Hours']} - {saidi_total:.6f} × {line_total_users}) / ({line_total_users} × {CONSTANTS['Annual_Power_Hours']}) × 100%")
    print(f"  ASAI = {asai:.6f} %")
    
    return {
        "线路类型": line_type,
        "总长度(km)": round(total_length, 4),
        "总用户数(台)": line_total_users,
        "总故障次数(次/年)": round(total_fault_count, 6),
        "总预安排次数(次/年)": round(total_scheduled_count, 6),
        "SAIDI-F": round(saidi_f, 6),
        "SAIDI-S": round(saidi_s, 6),
        "SAIDI合计": round(saidi_total, 6),
        "SAIFI-F": round(saifi_f, 6),
        "SAIFI-S": round(saifi_s, 6),
        "SAIFI合计": round(saifi_total, 6),
        "ASAI(%)": round(asai, 6)
    }

main_summary = calculate_summary(df_main_result, main_total_users, "主线")
branch_summary = calculate_summary(df_branch_result, branch_total_users, "分支")

# ==================== 第十步：全线路汇总（按技术方案5.2） ====================
print("\n" + "=" * 80)
print("【第十步】全线路汇总（按技术方案5.2加权平均）")
print("=" * 80)

print("\n全线路汇总计算过程:")
print(f"  全线路总长度 = {main_summary['总长度(km)']} + {branch_summary['总长度(km)']} = {main_summary['总长度(km)'] + branch_summary['总长度(km)']:.4f} km")
print(f"  全线路总用户数 = {main_total_users} + {branch_total_users} = {all_total_users} 台")

# 加权平均计算
all_saidi_f = (main_summary["SAIDI-F"] * main_total_users + branch_summary["SAIDI-F"] * branch_total_users) / all_total_users
all_saidi_s = (main_summary["SAIDI-S"] * main_total_users + branch_summary["SAIDI-S"] * branch_total_users) / all_total_users
all_saidi_total = all_saidi_f + all_saidi_s

all_saifi_f = (main_summary["SAIFI-F"] * main_total_users + branch_summary["SAIFI-F"] * branch_total_users) / all_total_users
all_saifi_s = (main_summary["SAIFI-S"] * main_total_users + branch_summary["SAIFI-S"] * branch_total_users) / all_total_users
all_saifi_total = all_saifi_f + all_saifi_s

print(f"\n  SAIDI-F加权 = ({main_summary['SAIDI-F']:.6f} × {main_total_users} + {branch_summary['SAIDI-F']:.6f} × {branch_total_users}) / {all_total_users}")
print(f"             = {all_saidi_f:.6f}")

print(f"  SAIDI-S加权 = ({main_summary['SAIDI-S']:.6f} × {main_total_users} + {branch_summary['SAIDI-S']:.6f} × {branch_total_users}) / {all_total_users}")
print(f"             = {all_saidi_s:.6f}")

print(f"  SAIDI合计 = {all_saidi_f:.6f} + {all_saidi_s:.6f} = {all_saidi_total:.6f}")

print(f"\n  SAIFI-F加权 = ({main_summary['SAIFI-F']:.6f} × {main_total_users} + {branch_summary['SAIFI-F']:.6f} × {branch_total_users}) / {all_total_users}")
print(f"             = {all_saifi_f:.6f}")

print(f"  SAIFI-S加权 = ({main_summary['SAIFI-S']:.6f} × {main_total_users} + {branch_summary['SAIFI-S']:.6f} × {branch_total_users}) / {all_total_users}")
print(f"             = {all_saifi_s:.6f}")

print(f"  SAIFI合计 = {all_saifi_f:.6f} + {all_saifi_s:.6f} = {all_saifi_total:.6f}")

# ASAI计算
all_theory_hours = all_total_users * CONSTANTS["Annual_Power_Hours"]
all_outage_hours = all_saidi_total * all_total_users
all_asai = ((all_theory_hours - all_outage_hours) / all_theory_hours) * 100

print(f"\n  ASAI计算:")
print(f"    全线路理论供电时间 = {all_total_users} × {CONSTANTS['Annual_Power_Hours']} = {all_theory_hours} 小时")
print(f"    全线路停电损失时间 = {all_saidi_total:.6f} × {all_total_users} = {all_outage_hours:.6f} 小时")
print(f"    ASAI = ({all_theory_hours} - {all_outage_hours:.6f}) / {all_theory_hours} × 100%")
print(f"         = {all_asai:.6f} %")

all_summary = {
    "线路类型": "全线路",
    "总长度(km)": round(main_summary["总长度(km)"] + branch_summary["总长度(km)"], 4),
    "总用户数(台)": all_total_users,
    "总故障次数(次/年)": round(main_summary["总故障次数(次/年)"] + branch_summary["总故障次数(次/年)"], 6),
    "总预安排次数(次/年)": round(main_summary["总预安排次数(次/年)"] + branch_summary["总预安排次数(次/年)"], 6),
    "SAIDI-F": round(all_saidi_f, 6),
    "SAIDI-S": round(all_saidi_s, 6),
    "SAIDI合计": round(all_saidi_total, 6),
    "SAIFI-F": round(all_saifi_f, 6),
    "SAIFI-S": round(all_saifi_s, 6),
    "SAIFI合计": round(all_saifi_total, 6),
    "ASAI(%)": round(all_asai, 6)
}

# ==================== 第十一步：汇总结果输出 ====================
print("\n" + "=" * 80)
print("【第十一步】最终汇总结果")
print("=" * 80)

summary_df = pd.DataFrame([main_summary, branch_summary, all_summary])
print("\n" + summary_df.to_string(index=False))

# ==================== 第十二步：输出到Excel ====================
print("\n" + "=" * 80)
print("【第十二步】输出到Excel文件")
print("=" * 80)

output_path = r"/mnt/d/pwkkx/10kV景704景水线_可靠性计算结果.xlsx"

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
wb.remove(wb.active)

# 主线分段明细
ws1 = wb.create_sheet(title="主线分段明细")
output_cols = ["分段编号", "长度(km)", "用户数(台)", "电缆权重", "架空权重", "敷设方式描述", "自动化状态", "故障率", "隔离时间",
               "故障次数(次/年)", "故障总时间(小时/次)", "预安排次数(次/年)",
               "SAIDI-F", "SAIDI-S", "SAIDI合计", "SAIFI-F", "SAIFI-S", "SAIFI合计"]
for r in dataframe_to_rows(df_main_result[output_cols], index=False, header=True):
    ws1.append(r)

# 分支分段明细
ws2 = wb.create_sheet(title="分支分段明细")
for r in dataframe_to_rows(df_branch_result[output_cols], index=False, header=True):
    ws2.append(r)

# 指标汇总
ws3 = wb.create_sheet(title="指标汇总")
for r in dataframe_to_rows(summary_df, index=False, header=True):
    ws3.append(r)

wb.save(output_path)
print(f"结果已保存到: {output_path}")

print("\n" + "=" * 80)
print("计算完成！")
print("=" * 80)
