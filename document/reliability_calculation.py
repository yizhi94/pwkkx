import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def reliability_calculation(excel_input_path, main_sheet_name, branch_sheet_name, excel_output_path):
    """
    适配你的Excel表头的供电可靠性指标计算脚本
    """
    # -------------------------- 1. 初始化基础参数（星能江夏数值） --------------------------
    param_config = {
        "fault_rate": {"电缆": 0.09282879, "混合": 0.108},
        "isolation_time": {True: 0.557, False: 2.0, "TRUE": 0.557, "FALSE": 2.0},
        "repair_time": 3.073,
        "scheduled_rate": 0.0221,
        "scheduled_total_time": 5.475,
        "annual_power_hours": 8760
    }

    # -------------------------- 2. 读取并校验输入Excel数据（适配你的表头） --------------------------
    try:
        main_df = pd.read_excel(excel_input_path, sheet_name=main_sheet_name)
        branch_df = pd.read_excel(excel_input_path, sheet_name=branch_sheet_name)
        print("✅ Excel文件读取成功")
    except Exception as e:
        print(f"❌ 读取Excel失败：{str(e)}")
        return

    # 【关键调整：适配你的Excel表头字段】
    # 主线Sheet字段映射：你的表头 → 脚本需要的字段
    main_df_rename = {
        "线路分段": "分段编号",
        "起点是否自动化": "自动化状态",
        "长度(km)": "长度(km)",
        "用户数量(台)": "用户数(台)",
        "线路型号": "敷设方式"
    }
    # 分支Sheet字段映射：你的表头 → 脚本需要的字段
    branch_df_rename = {
        "分支分段": "分段编号",
        "是否自动化": "自动化状态",
        "长度(km)": "长度(km)",
        "用户数量(台)": "用户数(台)",
        "线路型号": "敷设方式"
    }

    # 重命名字段并保留必要列
    main_df = main_df.rename(columns=main_df_rename)[list(main_df_rename.values())]
    branch_df = branch_df.rename(columns=branch_df_rename)[list(branch_df_rename.values())]

    # 校验必选字段
    required_cols = ["分段编号", "自动化状态", "长度(km)", "用户数(台)", "敷设方式"]
    for df, line_type in [(main_df, "主线"), (branch_df, "支线")]:
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            print(f"❌ {line_type}数据缺失字段：{', '.join(missing_cols)}")
            return
        # 清洗数据类型
        df["长度(km)"] = pd.to_numeric(df["长度(km)"], errors="coerce")
        df["用户数(台)"] = pd.to_numeric(df["用户数(台)"], errors="coerce")
        df = df[(df["长度(km)"] >= 0) & (df["用户数(台)"] >= 0)].dropna()
        if len(df) == 0:
            print(f"❌ {line_type}无有效数据")
            return

    # -------------------------- 3. 数据预处理（新增：从线路型号提取敷设方式） --------------------------
    def extract_laying_method(line_model):
        """从你的线路型号（如PD_YJV22）中提取敷设方式：YJV/YJLV为电缆，其他为混合"""
        line_model = str(line_model).upper()
        if any(keyword in line_model for keyword in ["YJV", "YJLV", "YJV22"]):
            return "电缆"
        else:
            return "混合"

    def preprocess_data(df, line_type):
        # 1. 从线路型号提取敷设方式
        df["敷设方式"] = df["敷设方式"].apply(extract_laying_method)
        # 2. 匹配故障率
        df["故障率"] = df["敷设方式"].map(param_config["fault_rate"])
        # 3. 统一自动化状态格式
        df["自动化状态"] = df["自动化状态"].apply(
            lambda x: x if isinstance(x, bool) else (True if str(x).upper() == "TRUE" else False)
        )
        df["隔离时间"] = df["自动化状态"].map(param_config["isolation_time"])
        # 4. 补充固定参数
        df["故障修复时间"] = param_config["repair_time"]
        df["预安排停电率"] = param_config["scheduled_rate"]
        df["预安排总时间"] = param_config["scheduled_total_time"]
        # 5. 标记有效分段
        df["有效分段"] = df["用户数(台)"] > 0
        # 6. 线路总用户数
        total_users = df["用户数(台)"].sum()
        print(f"✅ {line_type}预处理完成，总用户数：{total_users}台")
        return df, total_users

    main_df, main_total_users = preprocess_data(main_df, "主线")
    branch_df, branch_total_users = preprocess_data(branch_df, "支线")
    total_users_all = main_total_users + branch_total_users

    # -------------------------- 4. 分段级指标计算（逻辑不变，字段已适配） --------------------------
    def calculate_segment_indicators(df, line_total_users):
        # 故障相关指标
        df["故障次数(次/年)"] = np.where(
            df["有效分段"], df["长度(km)"] * df["故障率"], 0
        )
        df["故障总时间(小时/次)"] = df["隔离时间"] + df["故障修复时间"]
        df["SAIDI-F(小时/(户·年))"] = np.where(
            df["有效分段"] & (line_total_users > 0),
            (df["故障次数(次/年)"] * df["故障总时间(小时/次)"] * df["用户数(台)"]) / line_total_users,
            0
        )
        df["SAIFI-F(次/(户·年))"] = np.where(
            df["有效分段"] & (line_total_users > 0),
            (df["故障次数(次/年)"] * df["用户数(台)"]) / line_total_users,
            0
        )

        # 预安排相关指标
        df["预安排次数(次/年)"] = np.where(
            df["有效分段"], df["长度(km)"] * df["预安排停电率"], 0
        )
        df["SAIDI-S(小时/(户·年))"] = np.where(
            df["有效分段"] & (line_total_users > 0),
            (df["预安排次数(次/年)"] * df["预安排总时间"] * df["用户数(台)"]) / line_total_users,
            0
        )
        df["SAIFI-S(次/(户·年))"] = np.where(
            df["有效分段"] & (line_total_users > 0),
            (df["预安排次数(次/年)"] * df["用户数(台)"]) / line_total_users,
            0
        )

        # 合计指标
        df["SAIDI合计(小时/(户·年))"] = df["SAIDI-F(小时/(户·年))"] + df["SAIDI-S(小时/(户·年))"]
        df["SAIFI合计(次/(户·年))"] = df["SAIFI-F(次/(户·年))"] + df["SAIFI-S(次/(户·年))"]
        df[["SAIDI-F(小时/(户·年))", "SAIDI-S(小时/(户·年))", "SAIDI合计(小时/(户·年))",
            "SAIFI-F(次/(户·年))", "SAIFI-S(次/(户·年))", "SAIFI合计(次/(户·年))"]] = df[
            ["SAIDI-F(小时/(户·年))", "SAIDI-S(小时/(户·年))", "SAIDI合计(小时/(户·年))",
             "SAIFI-F(次/(户·年))", "SAIFI-S(次/(户·年))", "SAIFI合计(次/(户·年))"]].round(6)
        return df

    main_df = calculate_segment_indicators(main_df, main_total_users)
    branch_df = calculate_segment_indicators(branch_df, branch_total_users)
    print("✅ 分段级指标计算完成")

    # -------------------------- 5. 汇总级指标计算（逻辑不变） --------------------------
    def calculate_summary_indicators(df, line_total_users, line_type):
        total_length = df["长度(km)"].sum()
        total_fault_count = df["故障次数(次/年)"].sum()
        total_scheduled_count = df["预安排次数(次/年)"].sum()

        total_saidi_f = df["SAIDI-F(小时/(户·年))"].sum()
        total_saidi_s = df["SAIDI-S(小时/(户·年))"].sum()
        total_saidi = total_saidi_f + total_saidi_s
        total_saifi_f = df["SAIFI-F(次/(户·年))"].sum()
        total_saifi_s = df["SAIFI-S(次/(户·年))"].sum()
        total_saifi = total_saifi_f + total_saifi_s

        if line_total_users > 0:
            total_theory_hours = line_total_users * param_config["annual_power_hours"]
            total_outage_hours = total_saidi * line_total_users
            asai = ((total_theory_hours - total_outage_hours) / total_theory_hours) * 100
            asai = round(asai, 4)
        else:
            asai = 100.0

        summary = {
            "线路类型": line_type,
            "总长度(km)": round(total_length, 4),
            "总用户数(台)": line_total_users,
            "总故障次数(次/年)": round(total_fault_count, 6),
            "总预安排次数(次/年)": round(total_scheduled_count, 6),
            "SAIDI-F(小时/(户·年))": round(total_saidi_f, 6),
            "SAIDI-S(小时/(户·年))": round(total_saidi_s, 6),
            "SAIDI合计(小时/(户·年))": round(total_saidi, 6),
            "SAIFI-F(次/(户·年))": round(total_saifi_f, 6),
            "SAIFI-S(次/(户·年))": round(total_saifi_s, 6),
            "SAIFI合计(次/(户·年))": round(total_saifi, 6),
            "ASAI(%)": asai
        }
        return summary

    main_summary = calculate_summary_indicators(main_df, main_total_users, "主线")
    branch_summary = calculate_summary_indicators(branch_df, branch_total_users, "支线")

    # 全线路汇总
    all_summary = {
        "线路类型": "全线路",
        "总长度(km)": round(main_summary["总长度(km)"] + branch_summary["总长度(km)"], 4),
        "总用户数(台)": total_users_all,
        "总故障次数(次/年)": round(main_summary["总故障次数(次/年)"] + branch_summary["总故障次数(次/年)"], 6),
        "总预安排次数(次/年)": round(main_summary["总预安排次数(次/年)"] + branch_summary["总预安排次数(次/年)"], 6),
        "SAIDI-F(小时/(户·年))": round(
            (main_summary["SAIDI-F(小时/(户·年))"] * main_total_users + branch_summary[
                "SAIDI-F(小时/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "SAIDI-S(小时/(户·年))": round(
            (main_summary["SAIDI-S(小时/(户·年))"] * main_total_users + branch_summary[
                "SAIDI-S(小时/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "SAIDI合计(小时/(户·年))": round(
            (main_summary["SAIDI合计(小时/(户·年))"] * main_total_users + branch_summary[
                "SAIDI合计(小时/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "SAIFI-F(次/(户·年))": round(
            (main_summary["SAIFI-F(次/(户·年))"] * main_total_users + branch_summary[
                "SAIFI-F(次/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "SAIFI-S(次/(户·年))": round(
            (main_summary["SAIFI-S(次/(户·年))"] * main_total_users + branch_summary[
                "SAIFI-S(次/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "SAIFI合计(次/(户·年))": round(
            (main_summary["SAIFI合计(次/(户·年))"] * main_total_users + branch_summary[
                "SAIFI合计(次/(户·年))"] * branch_total_users) / total_users_all, 6
        ),
        "ASAI(%)": round(
            ((total_users_all * param_config["annual_power_hours"] -
              (main_summary["SAIDI合计(小时/(户·年))"] * main_total_users + branch_summary[
                  "SAIDI合计(小时/(户·年))"] * branch_total_users)) /
             (total_users_all * param_config["annual_power_hours"])) * 100, 4
        )
    }

    summary_df = pd.DataFrame([main_summary, branch_summary, all_summary])
    print("✅ 汇总级指标计算完成")

    # -------------------------- 6. 输出结果到Excel（保留你的原始分段信息） --------------------------
    try:
        wb = Workbook()
        wb.remove(wb.active)

        # 主线分段明细（保留你的原始字段+计算指标）
        ws1 = wb.create_sheet(title="主线分段明细")
        main_output_cols = [
            "分段编号", "自动化状态", "长度(km)", "用户数(台)", "敷设方式",
            "故障次数(次/年)", "故障总时间(小时/次)", "SAIDI-F(小时/(户·年))",
            "预安排次数(次/年)", "SAIDI-S(小时/(户·年))", "SAIDI合计(小时/(户·年))",
            "SAIFI-F(次/(户·年))", "SAIFI-S(次/(户·年))", "SAIFI合计(次/(户·年))"
        ]
        for r in dataframe_to_rows(main_df[main_output_cols], index=False, header=True):
            ws1.append(r)

        # 支线分段明细
        ws2 = wb.create_sheet(title="支线分段明细")
        branch_output_cols = main_output_cols
        for r in dataframe_to_rows(branch_df[branch_output_cols], index=False, header=True):
            ws2.append(r)

        # 指标汇总
        ws3 = wb.create_sheet(title="指标汇总")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws3.append(r)

        wb.save(excel_output_path)
        print(f"✅ 结果已输出到：{os.path.abspath(excel_output_path)}")
        print("\n📊 汇总结果预览：")
        print(summary_df.to_string(index=False))

    except Exception as e:
        print(f"❌ 输出Excel失败：{str(e)}")
        return

# -------------------------- 7. 脚本执行入口（用户需修改以下参数） --------------------------
if __name__ == "__main__":
    # 用户需根据实际情况修改以下3个参数
    INPUT_EXCEL_PATH =  r"D:\works\电网\人工智能\AI需求\配网全景拓扑\需求\技术调研\供电可靠性\算法\10kV安54新窑线.xlsx"  # 输入Excel文件路径
    MAIN_SHEET_NAME = "主线"               # 主线数据所在Sheet名称
    BRANCH_SHEET_NAME = "分支"            # 支线数据所在Sheet名称
    OUTPUT_EXCEL_PATH =  r"D:\works\电网\人工智能\AI需求\配网全景拓扑\需求\技术调研\供电可靠性\算法\10kV线路可靠性计算结果.xlsx"  # 输出Excel文件路径

    # 调用计算函数
    reliability_calculation(
        excel_input_path=INPUT_EXCEL_PATH,
        main_sheet_name=MAIN_SHEET_NAME,
        branch_sheet_name=BRANCH_SHEET_NAME,
        excel_output_path=OUTPUT_EXCEL_PATH
    )